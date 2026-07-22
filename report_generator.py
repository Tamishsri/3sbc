"""
report_generator.py
--------------------
Flattens the evaluated pipeline output and exports it to a formatted Excel report.

Output columns per row:
  - Team Name
  - Target Skill (AREA)
  - Target Location
  - Candidate Name/Title
  - Candidate LinkedIn URL
  - Match Score
  - AI Reasoning

Rows are sorted by Team Name (ASC) then Match Score (DESC).
Column headers are bold and columns are auto-fitted to content width.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_OUTPUT_FILENAME = "Sourced_Candidates_Report.xlsx"

OUTPUT_COLUMNS = [
    "Team Name",
    "Target Skill (AREA)",
    "Target Location",
    "Candidate Name/Title",
    "Candidate LinkedIn URL",
    "Match Score",
    "AI Reasoning",
]

# Header styling
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F4F8F")   # Dark navy
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)

# Score colour thresholds
SCORE_HIGH_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")   # Green  ≥ 75
SCORE_MED_FILL  = PatternFill(fill_type="solid", fgColor="FFEB9C")   # Yellow ≥ 50
SCORE_LOW_FILL  = PatternFill(fill_type="solid", fgColor="FFC7CE")   # Red    < 50

# Minimum column width (characters)
MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 60


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _flatten_data(evaluated_data: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Flatten the nested consultant → sourced_candidates structure into a
    flat list of rows, one per sourced candidate.
    """
    rows: list[dict[str, Any]] = []

    for consultant in evaluated_data:
        team_name = str(consultant.get("Team") or "").strip()
        target_skill = str(consultant.get("AREA") or "").strip()
        target_location = str(consultant.get("Location") or "").strip()

        candidates: list[dict[str, Any]] = consultant.get("sourced_candidates") or []

        if not candidates:
            # Still include a row so no consultant is silently lost
            rows.append(
                {
                    "Team Name": team_name,
                    "Target Skill (AREA)": target_skill,
                    "Target Location": target_location,
                    "Candidate Name/Title": "No candidates found",
                    "Candidate LinkedIn URL": "",
                    "Match Score": 0,
                    "AI Reasoning": "No candidates were sourced for this record.",
                }
            )
            continue

        for candidate in candidates:
            rows.append(
                {
                    "Team Name": team_name,
                    "Target Skill (AREA)": target_skill,
                    "Target Location": target_location,
                    "Candidate Name/Title": str(candidate.get("name_title") or ""),
                    "Candidate LinkedIn URL": str(candidate.get("linkedin_url") or ""),
                    "Match Score": int(candidate.get("match_score") or 0),
                    "AI Reasoning": str(candidate.get("reasoning") or ""),
                }
            )

    return rows


def _auto_fit_column_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Adjust each column's width to fit the longest cell value."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted = max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, max_length + 4))
        ws.column_dimensions[col_letter].width = adjusted


def _style_score_cell(cell: openpyxl.cell.cell.Cell, score: int) -> None:
    """Apply conditional colour formatting to a Match Score cell."""
    if score >= 75:
        cell.fill = SCORE_HIGH_FILL
    elif score >= 50:
        cell.fill = SCORE_MED_FILL
    else:
        cell.fill = SCORE_LOW_FILL

    cell.font = Font(bold=True, name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_report(
    evaluated_data: list[dict[str, Any]],
    output_filepath: str | Path = DEFAULT_OUTPUT_FILENAME,
) -> Path:
    """
    Generate a formatted Excel report from the evaluated pipeline output.

    Parameters
    ----------
    evaluated_data : list[dict]
        Output from ai_evaluator.evaluate_candidates() – fully nested
        structure with match_score and reasoning on each candidate.
    output_filepath : str or Path
        Where to save the Excel file. Defaults to 'Sourced_Candidates_Report.xlsx'
        in the current working directory.

    Returns
    -------
    Path
        Absolute path to the saved Excel file.
    """
    output_path = Path(output_filepath).resolve()
    print(f"[reporter] Generating report -> {output_path} ...")

    # 1. Flatten data
    rows = _flatten_data(evaluated_data)
    if not rows:
        print("[reporter] WARN: No rows to write -- evaluated_data is empty.")
        return output_path

    # 2. Build DataFrame
    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    # 3. Sort: Team Name ASC, Match Score DESC
    df.sort_values(
        by=["Team Name", "Match Score"],
        ascending=[True, False],
        inplace=True,
    )
    df.reset_index(drop=True, inplace=True)

    # 4. Write with pandas to get the initial file
    df.to_excel(str(output_path), index=False, engine="openpyxl")

    # Also export candidates_data.json for instant web app rendering
    json_path = output_path.parent / "candidates_data.json"
    try:
        json_records = df.to_dict(orient="records")
        with open(json_path, "w", encoding="utf-8") as jf:
            json.dump(json_records, jf, indent=2)
        print(f"[reporter] [OK] JSON data exported: {json_path}")
    except Exception as je:
        print(f"[reporter] WARN: Could not write candidates_data.json: {je}")

    # 5. Re-open with openpyxl for formatting
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active
    ws.title = "Candidate Matches"

    score_col_idx = OUTPUT_COLUMNS.index("Match Score") + 1  # 1-indexed

    # Style header row
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            # Alternate row shading for readability
            if row_idx % 2 == 0:
                if cell.fill.fgColor.rgb == "00000000":  # Only if not already coloured
                    cell.fill = PatternFill(fill_type="solid", fgColor="F0F4FF")

            # Wrap long text (AI Reasoning column)
            cell.alignment = Alignment(wrap_text=(col_idx == len(OUTPUT_COLUMNS)), vertical="top")

            # Colour-code the Match Score cell
            if col_idx == score_col_idx and isinstance(cell.value, (int, float)):
                _style_score_cell(cell, int(cell.value))

    # Auto-fit column widths
    _auto_fit_column_width(ws)

    # Set a sensible default row height for wrapped text rows
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30

    wb.save(str(output_path))
    print(f"[reporter] [OK] Report saved: {output_path}  ({len(df)} candidate rows)")
    return output_path


# ---------------------------------------------------------------------------
# Standalone test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # Build a minimal fake evaluated_data for formatting validation
    SAMPLE_DATA: list[dict[str, Any]] = [
        {
            "Team": "Team Alpha",
            "AREA": "SAP MM / Ariba",
            "Location": "USA",
            "NAME OF THE CONSULTANT": "John Doe",
            "sourced_candidates": [
                {
                    "name_title": "Alice Smith – SAP MM Consultant",
                    "linkedin_url": "https://www.linkedin.com/in/alice-smith",
                    "experience_summary": "SAP MM @ Accenture | Ariba Procurement @ Deloitte",
                    "skills": ["SAP MM", "Ariba", "S/4HANA"],
                    "match_score": 88,
                    "reasoning": "Strong SAP MM and Ariba background with S/4HANA exposure matching all key requirements.",
                },
                {
                    "name_title": "Bob Jones – ERP Specialist",
                    "linkedin_url": "https://www.linkedin.com/in/bob-jones",
                    "experience_summary": "SAP MM @ IBM",
                    "skills": ["SAP MM", "Materials Management"],
                    "match_score": 62,
                    "reasoning": "SAP MM experience is solid but lacks Ariba and S/4HANA skills required for this role.",
                },
            ],
        },
        {
            "Team": "Team Beta",
            "AREA": "Azure Data Engineer",
            "Location": "Canada",
            "NAME OF THE CONSULTANT": "Jane Roe",
            "sourced_candidates": [
                {
                    "name_title": "Carol White – Data Engineer",
                    "linkedin_url": "https://www.linkedin.com/in/carol-white",
                    "experience_summary": "Azure Data Engineer @ Microsoft",
                    "skills": ["Azure Data Factory", "Databricks", "PySpark"],
                    "match_score": 95,
                    "reasoning": "Excellent fit with direct Azure Data Engineering experience including ADF, Databricks, and PySpark.",
                },
            ],
        },
    ]

    out = generate_report(SAMPLE_DATA)
    print(f"\nReport written to: {out}")
